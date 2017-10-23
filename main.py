from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import bs4
import time
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import xlsxwriter
from openpyxl import load_workbook


wb = load_workbook(r'Input\Lenovo_PN.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
# shr-licznik dla wierszy w Excelu
shr = 2
pn = sheet['A'+str(shr)].value

while True:
    if sheet['A'+str(shr)].value is not None:
        print('Checking: '+'Sheet '+'A'+str(shr))
        print('PN: '+str(pn))

        driver = webdriver.Firefox()
        main_window = driver.current_window_handle
        driver.get("http://support.lenovo.com/us/pl/partslookup")
        elem = driver.find_element_by_id("partNumQuery")
        elem.clear()
        elem.send_keys(pn)
        elem.send_keys(Keys.RETURN)


        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="new-ResultsContent"]/div[1]/ul/li[2]/div[1]/span[6]')))
        driver.find_element_by_xpath('//*[@id="new-ResultsContent"]/div[1]/ul/li[2]/div[1]/span[6]').click()
        time.sleep(5)
        element = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, '//*[@id="new-ResultsContent"]/div[1]/ul/li[2]/div[4]/div/ul/li[1]/b[2]')))

        x = {}
        try:
            soup = bs4.BeautifulSoup(driver.page_source,'html.parser')
            x = {line.findAll('b',{'ng-bind' : 'Sub.PartNumber'})[0].getText(): line.findAll('b',{'ng-bind' : 'Sub.SubType'})[0].getText() for line in soup.findAll('li', {'ng-repeat' : 'Sub in parts.result.SubstitutionResult'})}
            for i in x :
                print(i)
        except:
            print("no results")

        workbook = xlsxwriter.Workbook(r'Output\Output.xlsx')
        worksheet = workbook.add_worksheet(str(pn))
        worksheet.set_column('A:C', 20)
        bold = workbook.add_format({'bold': True})

        worksheet.write('A1', 'Part#', bold)
        worksheet.write('B1', 'Sub Type', bold)
        r = 2
        for keys, values in x.items():
            worksheet.write('A'+str(r), keys)
            worksheet.write('B'+str(r), values)
            r += 1
        workbook.close()
        shr += 1
        print('Counter values is : '+str(shr))
        driver.switch_to_window(main_window)

print("done")




